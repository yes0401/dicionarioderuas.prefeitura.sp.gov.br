from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import sys
import json

def convert_xlsx_to_json(input_path, output_path):
	wb = load_workbook(input_path)

	rows = list(wb.active.iter_rows(values_only=True))
	headers = rows.pop(0)

	data = []
	for row in rows:
		entry = {}
		for idx, value in enumerate(row):
			if value is None:
				entry[headers[idx]] = None
			elif headers[idx] in ("created_at", "updated_at", "deleted_at"):
				entry[headers[idx]] = from_excel(value).strftime("%Y-%m-%dT%H:%M:%S.000000Z")
			else:
				entry[headers[idx]] = str(value)
		data.append(entry)

	with open(output_path, "w") as f:
		f.write(json.dumps(data))

if __name__ == "__main__":
	if len(sys.argv) < 3:
		print(f"Usage: {sys.argv[0]} input_path output_path")
		exit(1)

	input_path = sys.argv[1]
	output_path = sys.argv[2]

	convert_xlsx_to_json(input_path, output_path)
